"""
Bulk-import nhân viên từ file Excel (do phòng nhân sự điền).

Input  : /app/employees_import.xlsx     (copy vào container trước)
Output : /app/employees_import_result.xlsx  (chứa status + mật khẩu)

Quy tắc:
  - Idempotent: email đã tồn tại → bỏ qua (status="skipped").
  - Department/Custom role lookup theo TÊN (case-sensitive). Tên sai → row error.
  - Mật khẩu rỗng → auto-gen 10 ký tự ngẫu nhiên (urlsafe).
  - System role mặc định 'employee' nếu rỗng.
  - Commit từng row độc lập — 1 row lỗi không rollback các row khác.

Usage:
    docker cp ./tmp/employees_import.xlsx arkon_api:/app/
    docker exec arkon_api python -m app.scripts.import_employees
    docker cp arkon_api:/app/employees_import_result.xlsx ./tmp/
"""

import asyncio
import secrets
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select

from app.database import async_session_factory
from app.database.models import Department, Employee, Role
from app.services.auth_service import hash_password


INPUT_PATH = Path("/app/employees_import.xlsx")
OUTPUT_PATH = Path("/app/employees_import_result.xlsx")

# Column index (1-based) — must match generate_employees_template.py
COL_NAME = 1
COL_EMAIL = 2
COL_PASSWORD = 3
COL_DEPT = 4
COL_SYS_ROLE = 5
COL_CUSTOM_ROLE = 6
COL_NOTE = 7


def _gen_password() -> str:
    """10-char URL-safe random — collision-resistant + easy to share."""
    return secrets.token_urlsafe(8)[:10]


def _norm(s) -> str:
    return (str(s).strip() if s is not None else "")


async def main() -> None:
    if not INPUT_PATH.exists():
        print(f"✗ Không tìm thấy {INPUT_PATH}")
        print("  Copy vào container: docker cp ./tmp/employees_import.xlsx arkon_api:/app/")
        return

    wb_in = load_workbook(INPUT_PATH, data_only=True)
    if "Nhân viên" not in wb_in.sheetnames:
        print(f"✗ Sheet 'Nhân viên' không tồn tại trong {INPUT_PATH}")
        return
    ws_in = wb_in["Nhân viên"]

    # Read all data rows (skip header at row 1)
    rows: list[tuple[int, dict]] = []
    for r_idx in range(2, ws_in.max_row + 1):
        row = {
            "name": _norm(ws_in.cell(row=r_idx, column=COL_NAME).value),
            "email": _norm(ws_in.cell(row=r_idx, column=COL_EMAIL).value),
            "password": _norm(ws_in.cell(row=r_idx, column=COL_PASSWORD).value),
            "dept": _norm(ws_in.cell(row=r_idx, column=COL_DEPT).value),
            "sys_role": _norm(ws_in.cell(row=r_idx, column=COL_SYS_ROLE).value) or "employee",
            "custom_role": _norm(ws_in.cell(row=r_idx, column=COL_CUSTOM_ROLE).value),
            "note": _norm(ws_in.cell(row=r_idx, column=COL_NOTE).value),
        }
        # Skip blank rows
        if not row["name"] and not row["email"]:
            continue
        rows.append((r_idx, row))

    if not rows:
        print("Không có dòng nào để xử lý.")
        return

    print(f"Đang xử lý {len(rows)} dòng...\n")

    # Build lookup maps once
    async with async_session_factory() as session:
        dept_map = {
            d.name: d.id
            for d in (await session.execute(select(Department))).scalars().all()
        }
        role_map = {
            r.name: r.id
            for r in (await session.execute(select(Role))).scalars().all()
        }
        existing_emails = {
            e
            for e, in (await session.execute(select(Employee.email))).all()
        }

    # Per-row processing — independent commit
    results: list[dict] = []
    for r_idx, row in rows:
        out = {**row, "row": r_idx, "status": "", "message": "", "generated_password": ""}

        # Validate required
        if not row["name"]:
            out["status"] = "error"
            out["message"] = "Họ tên trống"
            results.append(out)
            continue
        if not row["email"]:
            out["status"] = "error"
            out["message"] = "Email trống"
            results.append(out)
            continue
        if "@" not in row["email"]:
            out["status"] = "error"
            out["message"] = "Email không đúng định dạng"
            results.append(out)
            continue
        if not row["dept"]:
            out["status"] = "error"
            out["message"] = "Phòng ban trống"
            results.append(out)
            continue

        # Duplicate check
        if row["email"].lower() in {e.lower() for e in existing_emails}:
            out["status"] = "skipped"
            out["message"] = "Email đã tồn tại"
            results.append(out)
            continue

        # Department lookup
        dept_id = dept_map.get(row["dept"])
        if dept_id is None:
            out["status"] = "error"
            out["message"] = f"Phòng ban không tồn tại: '{row['dept']}'"
            results.append(out)
            continue

        # Custom role lookup (optional)
        custom_role_id = None
        if row["custom_role"]:
            custom_role_id = role_map.get(row["custom_role"])
            if custom_role_id is None:
                out["status"] = "error"
                out["message"] = f"Vai trò không tồn tại: '{row['custom_role']}'"
                results.append(out)
                continue

        # System role validation
        if row["sys_role"] not in ("admin", "employee"):
            out["status"] = "error"
            out["message"] = f"Vai trò hệ thống không hợp lệ: '{row['sys_role']}' (chỉ admin/employee)"
            results.append(out)
            continue

        # Password (gen if blank)
        pw = row["password"]
        if not pw:
            pw = _gen_password()
            out["generated_password"] = pw
        elif len(pw) < 8:
            out["status"] = "error"
            out["message"] = "Mật khẩu < 8 ký tự"
            results.append(out)
            continue
        else:
            out["generated_password"] = pw

        # Create — independent transaction per row
        try:
            async with async_session_factory() as session:
                emp = Employee(
                    name=row["name"],
                    email=row["email"],
                    password_hash=hash_password(pw),
                    role=row["sys_role"],
                    department_id=dept_id,
                    custom_role_id=custom_role_id,
                    is_active=True,
                )
                session.add(emp)
                await session.commit()
                await session.refresh(emp)
                out["status"] = "created"
                out["message"] = f"id={emp.id}"
                existing_emails.add(row["email"])  # Prevent duplicate within same file
        except Exception as e:
            out["status"] = "error"
            out["message"] = f"DB error: {type(e).__name__}: {e}"

        results.append(out)
        icon = {"created": "✓", "skipped": "⊘", "error": "✗"}.get(out["status"], "?")
        print(f"  {icon} [{r_idx:3d}] {row['email']:35s} → {out['status']:8s} {out['message'][:60]}")

    # Summary
    by_status = {"created": 0, "skipped": 0, "error": 0}
    for r in results:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
    print(f"\nTổng: {len(results)}  ✓ {by_status['created']} tạo  ⊘ {by_status['skipped']} bỏ qua  ✗ {by_status['error']} lỗi")

    # Write result workbook
    wb_out = Workbook()
    ws = wb_out.active
    ws.title = "Kết quả"
    headers = [
        "Dòng", "Họ tên", "Email", "Mật khẩu (chia sẻ qua kênh bảo mật)",
        "Phòng ban", "Vai trò hệ thống", "Vai trò tùy chỉnh",
        "Trạng thái", "Chi tiết",
    ]
    ws.append(headers)
    widths = [6, 26, 32, 24, 32, 18, 30, 12, 60]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + c - 1)].width = w
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    status_fill = {
        "created": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "skipped": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "error":   PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    }
    for r in results:
        ws.append([
            r["row"], r["name"], r["email"],
            r["generated_password"],
            r["dept"], r["sys_role"], r["custom_role"],
            r["status"], r["message"],
        ])
        last = ws.max_row
        fill = status_fill.get(r["status"])
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(row=last, column=c).fill = fill

    wb_out.save(OUTPUT_PATH)
    print(f"\n✓ Kết quả ghi ra: {OUTPUT_PATH}")
    print(f"  → Copy ra host: docker cp arkon_api:{OUTPUT_PATH} ./tmp/")
    print(f"  ⚠ File chứa mật khẩu plaintext — KHÔNG share rộng, xóa sau khi đã chuyển credentials.")


if __name__ == "__main__":
    asyncio.run(main())
