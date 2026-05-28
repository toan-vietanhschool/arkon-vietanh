"""
Tạo file Excel template để phòng nhân sự nhập danh sách nhân viên.

Output: D:\\arkon-vietanh\\tmp\\employees_import_template.xlsx (hoặc cwd nếu
chạy trong container thì /app/tmp/ — copy ra sau).

Cấu trúc:
  - Sheet 'Nhân viên'       — danh sách điền (header + 3 ví dụ + 47 dòng trống)
  - Sheet 'Hướng dẫn'       — mô tả cột + quy tắc validation
  - Sheet 'Phòng ban'       — danh mục phòng ban hợp lệ (lấy từ DB live)
  - Sheet 'Vai trò'         — danh mục custom role hợp lệ (lấy từ DB live)

Cột bắt buộc: name, email, department
Cột optional: password (auto-gen nếu rỗng), custom_role (mặc định = không có)
              role (admin/employee — mặc định employee)

Usage (chạy trong container):
    docker exec arkon_api python -m app.scripts.generate_employees_template
    docker cp arkon_api:/app/employees_import_template.xlsx ./tmp/
"""

import asyncio
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select

from app.database import async_session_factory
from app.database.models import Department, Role


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
NOTE_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

COLUMNS: list[tuple[str, int, str]] = [
    # (header, width, comment)
    ("Họ tên *", 28, "Bắt buộc — tên đầy đủ"),
    ("Email *", 32, "Bắt buộc — phải duy nhất trong hệ thống"),
    ("Mật khẩu tạm thời", 22, "Để trống → script auto-gen ngẫu nhiên 10 ký tự"),
    ("Phòng ban *", 36, "Bắt buộc — chọn từ dropdown (sheet Phòng ban)"),
    ("Vai trò hệ thống", 18, "admin / employee (mặc định: employee)"),
    ("Vai trò tùy chỉnh", 32, "Chọn từ dropdown (sheet Vai trò) — để trống nếu không gán"),
    ("Ghi chú", 30, "Tùy chọn — context cho admin"),
]


async def _fetch_lookup_data():
    async with async_session_factory() as session:
        depts = (await session.execute(select(Department).order_by(Department.name))).scalars().all()
        roles = (await session.execute(select(Role).order_by(Role.is_system.desc(), Role.name))).scalars().all()
    return (
        [(d.name, (d.description or "")[:120]) for d in depts],
        [(r.name, r.is_system, (r.description or "")[:120]) for r in roles],
    )


def _style_header_row(ws, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"


def _build_employees_sheet(wb: Workbook, depts: list[tuple[str, str]], roles: list[tuple[str, bool, str]]) -> None:
    ws = wb.active
    ws.title = "Nhân viên"

    # Header
    for c, (header, width, _) in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=c, value=header)
        ws.column_dimensions[get_column_letter(c)].width = width
    _style_header_row(ws, len(COLUMNS))

    # Ví dụ rows
    examples = [
        ("Nguyễn Văn An", "nv.an@vietanh.edu.vn", "", "Phòng Tuyển Sinh", "employee", "Chuyên viên hành chính", "Tuyển sinh đợt T6/2026"),
        ("Trần Thị Bình", "tt.binh@vietanh.edu.vn", "TempPass!2026", "Khối Tiểu Học", "employee", "Giáo viên", "GVCN lớp 3A"),
        ("Lê Hoàng Cường", "lh.cuong@vietanh.edu.vn", "", "Phòng Công Nghệ Thông Tin", "admin", "", "IT Manager — admin hệ thống"),
    ]
    for r_idx, row in enumerate(examples, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(vertical="center")
            if r_idx <= 4:
                cell.fill = NOTE_FILL

    # Data validation dropdowns
    dept_names = [d[0] for d in depts]
    role_names = [r[0] for r in roles]

    # Dept dropdown (col D), 1000 rows
    if dept_names:
        dv_dept = DataValidation(
            type="list",
            formula1=f'"{",".join(dept_names)}"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Phòng ban không hợp lệ",
            error="Chọn từ danh sách trong sheet 'Phòng ban'",
        )
        ws.add_data_validation(dv_dept)
        dv_dept.add(f"D2:D1001")

    # System role dropdown (col E)
    dv_sys_role = DataValidation(
        type="list",
        formula1='"admin,employee"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Vai trò hệ thống không hợp lệ",
        error="Chọn 'admin' hoặc 'employee' (hoặc để trống = employee)",
    )
    ws.add_data_validation(dv_sys_role)
    dv_sys_role.add("E2:E1001")

    # Custom role dropdown (col F)
    if role_names:
        # Excel cell limit ~255 chars for inline list — use sheet reference if long
        if sum(len(n) for n in role_names) + len(role_names) > 240:
            dv_role = DataValidation(
                type="list",
                formula1=f"=Vai_tro!$A$2:$A${len(role_names) + 1}",
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Vai trò không hợp lệ",
                error="Chọn từ danh sách trong sheet 'Vai trò'",
            )
        else:
            dv_role = DataValidation(
                type="list",
                formula1=f'"{",".join(role_names)}"',
                allow_blank=True,
            )
        ws.add_data_validation(dv_role)
        dv_role.add("F2:F1001")


def _build_instruction_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Hướng dẫn")
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 100

    rows = [
        ("", "HƯỚNG DẪN ĐIỀN FILE EMPLOYEE IMPORT", True),
        ("", "", False),
        ("1.", "Sheet 'Nhân viên' — mỗi dòng 1 nhân viên cần tạo tài khoản.", False),
        ("2.", "Cột có dấu (*) là BẮT BUỘC: Họ tên, Email, Phòng ban.", False),
        ("3.", "Email phải duy nhất. Nếu đã tồn tại trong hệ thống, script sẽ BỎ QUA dòng đó.", False),
        ("4.", "Mật khẩu tạm thời: để trống thì script sẽ auto-gen 10 ký tự ngẫu nhiên.", False),
        ("5.", "Mật khẩu nhập thủ công phải ≥ 8 ký tự (theo chính sách hệ thống).", False),
        ("6.", "Phòng ban: chọn từ dropdown. Tên phải khớp CHÍNH XÁC với sheet 'Phòng ban'.", False),
        ("7.", "Vai trò hệ thống: 'admin' = quản trị viên toàn hệ thống (bypass mọi RBAC).", False),
        ("", "                     'employee' = nhân viên thường (default, RBAC theo Vai trò tùy chỉnh).", False),
        ("8.", "Vai trò tùy chỉnh: gán role custom (vd Giáo viên, Trưởng phòng, Hiệu trưởng).", False),
        ("", "                     Để trống → nhân viên dùng EMPLOYEE_DEFAULT_PERMISSIONS (6 quyền baseline).", False),
        ("9.", "Sau khi điền xong, lưu file với tên: employees_import.xlsx", False),
        ("10.", "Đưa file cho admin chạy: docker exec arkon_api python -m app.scripts.import_employees", False),
        ("", "", False),
        ("", "FILE KẾT QUẢ", True),
        ("", "Script chạy xong sẽ tạo file employees_import_result.xlsx chứa:", False),
        ("", "  - Trạng thái mỗi dòng: ✓ Đã tạo / ⊘ Trùng email / ✗ Lỗi (với lý do)", False),
        ("", "  - Mật khẩu tạm thời (kể cả khi auto-gen) — chuyển cho từng nhân viên qua kênh bảo mật", False),
        ("", "  - Nhân viên login lần đầu nên đổi mật khẩu ngay", False),
        ("", "", False),
        ("", "VÍ DỤ", True),
        ("", "3 dòng đầu trong sheet 'Nhân viên' là VÍ DỤ — xóa hoặc sửa trước khi import.", False),
        ("", "", False),
        ("", "BẢO MẬT", True),
        ("", "File employees_import_result.xlsx chứa mật khẩu plaintext — KHÔNG share rộng,", False),
        ("", "KHÔNG commit vào git. Xóa sau khi đã chuyển credentials cho từng người.", False),
    ]

    for r_idx, (col_a, col_b, is_heading) in enumerate(rows, start=1):
        ws.cell(row=r_idx, column=1, value=col_a).alignment = Alignment(vertical="top")
        cell_b = ws.cell(row=r_idx, column=2, value=col_b)
        cell_b.alignment = Alignment(vertical="top", wrap_text=True)
        if is_heading:
            cell_b.font = Font(bold=True, size=12, color="2563EB")


def _build_lookup_sheets(wb: Workbook, depts: list[tuple[str, str]], roles: list[tuple[str, bool, str]]) -> None:
    # Phòng ban
    ws_d = wb.create_sheet("Phòng ban")
    ws_d.append(["Tên phòng ban (dùng giá trị này)", "Mô tả"])
    ws_d.column_dimensions["A"].width = 36
    ws_d.column_dimensions["B"].width = 90
    for name, desc in depts:
        ws_d.append([name, desc])
    _style_header_row(ws_d, 2)

    # Vai trò (referenced internally as Vai_tro for formula compat)
    ws_r = wb.create_sheet("Vai_tro")
    ws_r.append(["Tên vai trò (dùng giá trị này)", "Hệ thống?", "Mô tả"])
    ws_r.column_dimensions["A"].width = 36
    ws_r.column_dimensions["B"].width = 12
    ws_r.column_dimensions["C"].width = 80
    for name, is_system, desc in roles:
        ws_r.append([name, "✓" if is_system else "", desc])
    _style_header_row(ws_r, 3)
    # Hide spaces in title for user, friendly title via tab color
    ws_r.sheet_properties.tabColor = "10B981"


async def main() -> None:
    depts, roles = await _fetch_lookup_data()
    print(f"  Đã đọc {len(depts)} phòng ban, {len(roles)} vai trò")

    wb = Workbook()
    _build_employees_sheet(wb, depts, roles)
    _build_instruction_sheet(wb)
    _build_lookup_sheets(wb, depts, roles)

    out_path = Path("/app/employees_import_template.xlsx")
    wb.save(out_path)
    print(f"\n✓ Đã tạo: {out_path}")
    print(f"  → Copy ra host: docker cp arkon_api:{out_path} ./tmp/")


if __name__ == "__main__":
    asyncio.run(main())
